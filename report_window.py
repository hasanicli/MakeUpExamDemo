import tkinter as tk
import tkinter.ttk as ttk
from tkinter import messagebox
from tkinter.filedialog import asksaveasfile
from data import Data
from json_file import JsonFile
from openpyxl import Workbook
from openpyxl.styles import alignment, Font
from teacher import Teacher


class ReportWindow(tk.Toplevel):
    def __init__(self):
        super().__init__()

        self.geometry("320x240")
        self.resizable = False
        self.title("Gözlem")
        self.grab_set()

        self._create_widgets()

    def save(self, name):
        files = [('Excel Document', '*.xlsx'), ('All Files', '*.*')]
        file = asksaveasfile(filetypes=files, defaultextension='*.xlsx', parent=self, initialfile=name)
        if file:
            return file.name
        return None

    @staticmethod
    def _center_text_to_cell(sht, pairs):
        for pair in pairs:
            rw, cl = pair
            sht.cell(row=rw, column=cl).alignment = alignment.Alignment(horizontal='center', vertical='center')

    def _create_sheet(self):
        data_control = Data("data.json")
        teacher_control = Teacher()

        lessons = data_control.get_all_lessons()
        students = data_control.get_student_numbers_with_names()
        lessons_with_numbers = data_control.get_lessons_with_numbers()
        teachers = teacher_control.get_branch_and_name()
        dates = data_control.get_all_date_and_time()

        wb = Workbook()

        ws1 = wb.active
        ws1.title = "Öğrenci-Ders"

        ws2 = wb.create_sheet("Öğretmen-Ders")
        ws3 = wb.create_sheet("Tarih-Ders")
        ws4 = wb.create_sheet("Öğretmen-Tarih")
        ws5 = wb.create_sheet("Öğrenci-Tarih")

        # ws1
        lc = self._fill_columns(lessons, ws1, 1, 2)
        lr = self._fill_rows(students, ws1, 1, 2)

        for row in range(2, lr + 1):
            student_number = ws1.cell(row=row, column=1).value.split("_")[0]
            for col in range(2, lc + 1):
                lesson_name = ws1.cell(row=1, column=col).value
                if student_number in lessons_with_numbers[lesson_name]:
                    ws1.cell(row=row, column=col).value = "X"
                    ws1.cell(row=row, column=col).alignment = alignment.Alignment(horizontal='center',
                                                                                  vertical='center')
                    ws1.cell(row=row, column=col).font = Font(bold=True)
        ws1.freeze_panes = ws1["B2"]

        # ws2
        lc = self._fill_columns(lessons, ws2, 1, 2)
        lr = self._fill_rows(teachers, ws2, 1, 2)
        for row in range(2, lr + 1):
            teacher_name = ws2.cell(row=row, column=1).value.split("_")[1]
            for col in range(2, lc + 1):
                lesson_name = ws2.cell(row=1, column=col).value
                if teacher_name in data_control.get_teachers_of_lesson(lesson_name):
                    ws2.cell(row=row, column=col).value = "X"
                    ws2.cell(row=row, column=col).alignment = alignment.Alignment(horizontal='center',
                                                                                  vertical='center')
                    ws2.cell(row=row, column=col).font = Font(bold=True)
        ws2.freeze_panes = ws2["B2"]

        # ws3
        lc = self._fill_columns(lessons, ws3, 1, 2)
        lr = self._fill_rows(dates, ws3, 1, 2)
        for row in range(2, lr + 1):
            date_and_time = ws3.cell(row=row, column=1).value
            for col in range(2, lc + 1):
                lesson_name = ws3.cell(row=1, column=col).value
                if date_and_time == data_control.get_date_and_time(lesson_name):
                    ws3.cell(row=row, column=col).value = "X"
                    ws3.cell(row=row, column=col).alignment = alignment.Alignment(horizontal='center',
                                                                                  vertical='center')
                    ws3.cell(row=row, column=col).font = Font(bold=True)
        ws3.freeze_panes = ws3["B2"]

        # ws4
        lc = self._fill_columns(teachers, ws4, 1, 2)
        lr = self._fill_rows(dates, ws4, 1, 2)
        for row in range(2, lr + 1):
            date_and_time = ws4.cell(row=row, column=1).value
            for col in range(2, lc + 1):
                teacher_name = ws4.cell(row=1, column=col).value.split("_")[1]
                if date_and_time in data_control.get_duty_dates_for_teachers(teacher_name):
                    ws4.cell(row=row, column=col).value = "X"
                    ws4.cell(row=row, column=col).alignment = alignment.Alignment(horizontal='center',
                                                                                  vertical='center')
                    ws4.cell(row=row, column=col).font = Font(bold=True)
        ws4.freeze_panes = ws4["B2"]

        # ws5
        lc = self._fill_columns(students, ws5, 1, 2)
        lr = self._fill_rows(dates, ws5, 1, 2)
        for row in range(2, lr + 1):
            date_and_time = ws5.cell(row=row, column=1).value
            for col in range(2, lc + 1):
                student_number = ws5.cell(row=1, column=col).value.split("_")[0]
                if date_and_time in data_control.get_student_dates(student_number):
                    ws5.cell(row=row, column=col).value = "X"
                    ws5.cell(row=row, column=col).alignment = alignment.Alignment(horizontal='center',
                                                                                  vertical='center')
                    ws5.cell(row=row, column=col).font = Font(bold=True)
        ws5.freeze_panes = ws5["B2"]

        try:
            sht_name = 'Çarşaflar.xlsx'
            full_address = self.save(sht_name)
            if full_address:
                wb.save(full_address)
        except PermissionError:
            tk.messagebox.showerror("Hata", "Kaydetmek istediğininz dosya açık", parent=self)

    @staticmethod
    def _fill_columns(items, sht_name, row, starting):
        for index, lesson in enumerate(items, starting):
            sht_name.cell(row=row, column=index, value=lesson)
            sht_name.cell(row=row, column=index).alignment = alignment.Alignment(horizontal='center', vertical='bottom',
                                                                                 text_rotation=90)
            sht_name.cell(row=1, column=index).font = Font(bold=True)
            col_letter = sht_name.cell(row=1, column=index).column_letter
            sht_name.column_dimensions[col_letter].width = 3
        return sht_name.max_column

    def _fill_rows(self, items, sht_name, column, starting):
        for i, student in enumerate(items, starting):
            sht_name.cell(row=i, column=column, value=student)
            sht_name.cell(row=i, column=column, value=student).font = Font(bold=True)

        self._adjust_column_width_autosize(sht_name, "A")
        return sht_name.max_row

    @staticmethod
    def _adjust_column_width_autosize(ws, letter):
        """https://stackoverflow.com/questions/39529662/python-automatically-adjust-width-of-an-excel-files-columns"""
        max_length = 0
        for cell in ws[letter]:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass

        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[letter].width = adjusted_width

    def _create_daily_signature_list(self):
        data_control = Data("data.json")
        fixed_info = JsonFile.read("fixed_info.json")

        # lessons = data_control.get_all_lessons()
        dates = sorted(list(set([i.split()[0] for i in data_control.get_all_date_and_time()])))

        wb = Workbook()
        for i in dates:
            new_name = i.replace("/", ".")
            wb.create_sheet(new_name)
            ws = wb[new_name]
            ws.merge_cells("A1:M1")
            ws.merge_cells("A2:M2")
            ws.merge_cells("A3:M3")
            ws.merge_cells("A4:M4")
            ws.merge_cells("A5:M5")

            if fixed_info:
                if fixed_info["county"] == "MERKEZ":
                    ws.cell(row=1, column=1, value=fixed_info["city"] + " VALİLİĞİ")
                    ws.cell(row=2, column=1, value=fixed_info["school_name"])
                    ws.cell(row=3, column=1, value=fixed_info["exam_term_name"] + " SORUMLULUK SINAVLARI")
                    ws.cell(row=4, column=1, value=i)
                else:
                    ws.cell(row=1, column=1, value=fixed_info["county"] + " KAYMKAMLIĞI")
                    ws.cell(row=2, column=1, value=fixed_info["school_name"])
                    ws.cell(row=3, column=1, value=fixed_info["exam_term_name"] + " SORUMLULUK SINAVLARI")
                    ws.cell(row=4, column=1, value=i)
                self._center_text_to_cell(ws, [(1, 1), (1, 2), (1, 3), (1, 4)])

            ws.cell(row=6, column=1).value = "Sıra"

            ws.merge_cells("B6:C6")
            ws.cell(row=6, column=2).value = "Öğretmen"

            ws.merge_cells("D6:J6")
            ws.cell(row=6, column=4).value = "Ders Adı"

            ws.cell(row=6, column=11).value = "Saat"

            ws.merge_cells("L6:M6")
            ws.cell(row=6, column=12).value = "İmza"

            ds = data_control.get_dataset_of_date(i)

            rw = 7
            for lesson, hour in ds:
                lesson_teachers = data_control.get_teachers_of_lesson(lesson)
                for teacher in lesson_teachers:
                    ws.cell(row=rw, column=1).value = rw - 6

                    ws.merge_cells("B" + str(rw) + ":C" + str(rw))
                    ws.cell(row=rw, column=2).value = teacher

                    ws.merge_cells("D" + str(rw) + ":J" + str(rw))
                    ws.cell(row=rw, column=4).value = lesson

                    ws.cell(row=rw, column=11).value = hour

                    ws.merge_cells("L" + str(rw) + ":M" + str(rw))
                    rw += 1

            self.set_border(ws, "A" + str(6) + ":M" + str(rw - 1))

            ws.merge_cells("D" + str(rw + 3) + ":J" + str(rw + 3))
            ws.merge_cells("D" + str(rw + 4) + ":J" + str(rw + 4))
            ws.merge_cells("D" + str(rw + 5) + ":J" + str(rw + 5))
            ws.merge_cells("D" + str(rw + 6) + ":J" + str(rw + 6))

            ws.cell(row=rw + 3, column=4, value=i)
            ws.cell(row=rw + 5, column=4, value=fixed_info["headmaster_name"])
            ws.cell(row=rw + 6, column=4, value="Okul Müdürü")
            self._center_text_to_cell(ws, [(rw + 3, 4), (rw + 5, 4), (rw + 6, 4)])

            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.print_area = "A1:M32"
        else:
            del wb["Sheet"]

        try:
            sht_name = 'İmza Sirküsü.xlsx'
            full_address = self.save(sht_name)
            if full_address:
                wb.save(full_address)
        except PermissionError:
            tk.messagebox.showerror("Hata", "Kaydetmek istediğininz dosya açık", parent=self)

    @staticmethod
    def set_border(ws, cell_range):
        from openpyxl.styles import Border, Side
        thin = Side(border_style="thin", color="000000")
        for row in ws[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    def _create_teacher_tasks_list(self):
        """Ogretmen gorevleri"""
        data_control = Data("data.json")
        teachers = JsonFile.read("teachers.json").keys()

        wb = Workbook()
        for teacher in teachers:
            teacher_tasks_list = data_control.get_task_dates_and_tasks_for_teachers(teacher)
            ws = wb.create_sheet(teacher)
            self._create_header_vertical(ws, teacher)

            ws.merge_cells("B15:G15")
            ws.merge_cells("H15:I15")
            ws.cell(row=15, column=1, value="Sıra")
            ws.cell(row=15, column=2, value="Sınav adı")
            ws.cell(row=15, column=8, value="Sınav tarihi")
            for index, task in enumerate(teacher_tasks_list):
                ws.cell(row=index + 16, column=1, value=index + 1)

                ws.merge_cells("B" + str(index + 16) + ":G" + str(index + 16))
                ws.cell(row=index + 16, column=2, value=task[0])

                ws.merge_cells("H" + str(index + 16) + ":I" + str(index + 16))
                ws.cell(row=index + 16, column=8, value=task[1])
                self.set_border(ws, "A" + str(15) + ":I" + str(index + 16))
        else:
            del wb["Sheet"]

        try:
            sht_name = 'Öğretmen Görev.xlsx'
            full_address = self.save(sht_name)
            if full_address:
                wb.save(full_address)
        except PermissionError:
            tk.messagebox.showerror("Hata", "Kaydetmek istediğininz dosya açık", parent=self)

    def _create_header_vertical(self, sht, teacher):
        fixed_info = JsonFile.read("fixed_info.json")

        sht.merge_cells("A1:I1")
        sht.merge_cells("A2:I2")
        sht.merge_cells("A3:I3")

        if fixed_info:
            if fixed_info["county"] == "MERKEZ":
                sht.cell(row=1, column=1, value=fixed_info["city"] + " VALİLİĞİ")
            else:
                sht.cell(row=1, column=1, value=fixed_info["county"] + " KAYMKAMLIĞI")
            sht.cell(row=2, column=1, value=fixed_info["school_name"])
            sht.cell(row=3, column=1, value=fixed_info["exam_term_name"] + " SORUMLULUK SINAVLARI")

            self._center_text_to_cell(sht, [(1, 1), (2, 1), (3, 1)])

            sht.merge_cells("B5:C5")
            sht.cell(row=5, column=1, value="Tarih : ")
            sht.cell(row=5, column=2, value=fixed_info["document_date"])

            sht.merge_cells("B6:C6")
            sht.cell(row=6, column=1, value="Sayı : ")
            sht.cell(row=6, column=2, value=fixed_info["document_number"])

            sht.merge_cells("A8:I8")
            sht.merge_cells("A9:I9")
            sht.merge_cells("A10:I10")

            txt1 = (
                f"    Sayın {teacher}, {fixed_info["exam_term_name"]} dönemine ait görevleriniz aşağıya çıkarılmıştır.")
            txt2 = f"Görevlerinizi titizlikle yerine getirmeniz hususunda gereğini;"
            txt3 = f"     Rica ederim."
            sht.cell(row=8, column=1, value=txt1)
            sht.cell(row=9, column=1, value=txt2)
            sht.cell(row=10, column=1, value=txt3)
            sht.merge_cells("H11:I11")
            sht.merge_cells("H12:I12")
            sht.merge_cells("H13:I13")

            sht.cell(row=11, column=8, value=fixed_info["document_date"])
            sht.cell(row=12, column=8, value=fixed_info["headmaster_name"])
            sht.cell(row=13, column=8, value="Okul Müdürü")

            self._center_text_to_cell(sht, [(11, 8), (12, 8), (13, 8)])

    def _create_consumable_list(self, lesson):
        """Sarf tutanagi"""

    def _create_notes_list(self, lesson):
        """Not cizelgesi"""
        pass

    def _create_exam_list(self):
        """Aski listesi"""
        data_control = Data("data.json")
        fixed_info = JsonFile.read("fixed_info.json")

        lessons = data_control.get_all_lessons()
        # dates = sorted(list(set([i.split()[0] for i in data_control.get_all_date_and_time()])))

        wb = Workbook()
        ws = wb.active
        ws.title = "Sınav listesi"

        ws.merge_cells("A1:I1")
        ws.merge_cells("A2:I2")
        ws.merge_cells("A3:I3")
        ws.merge_cells("A4:I4")

        if fixed_info:
            if fixed_info["county"] == "MERKEZ":
                ws.cell(row=1, column=1, value=fixed_info["city"] + " VALİLİĞİ")
                ws.cell(row=2, column=1, value=fixed_info["school_name"])
                ws.cell(row=3, column=1, value=fixed_info["exam_term_name"] + " SORUMLULUK SINAVLARI")
            else:
                ws.cell(row=1, column=1, value=fixed_info["county"] + " KAYMKAMLIĞI")
                ws.cell(row=2, column=1, value=fixed_info["school_name"])
                ws.cell(row=3, column=1, value=fixed_info["exam_term_name"] + " SORUMLULUK SINAVLARI")
            ws.cell(row=1, column=1).alignment = alignment.Alignment(horizontal='center', vertical='center')
            ws.cell(row=2, column=1).alignment = alignment.Alignment(horizontal='center', vertical='center')
            ws.cell(row=3, column=1).alignment = alignment.Alignment(horizontal='center', vertical='center')

        ws.cell(row=5, column=1).value = "Sıra"

        ws.merge_cells("B5:G5")
        ws.cell(row=5, column=2).value = "Dersin adı"

        ws.merge_cells("H5:I5")
        ws.cell(row=5, column=8).value = "Tarih-saat"

        rw = 6
        for lesson in lessons:
            lesson_date_and_time = data_control.get_date_and_time(lesson)

            ws.cell(row=rw, column=1).value = rw - 5

            ws.merge_cells("B" + str(rw) + ":G" + str(rw))
            ws.cell(row=rw, column=2).value = lesson

            ws.merge_cells("H" + str(rw) + ":I" + str(rw))
            ws.cell(row=rw, column=8).value = lesson_date_and_time

            rw += 1
        self.set_border(ws, "A" + str(5) + ":I" + str(rw - 1))

        ws.merge_cells("C" + str(rw + 3) + ":G" + str(rw + 3))
        ws.merge_cells("C" + str(rw + 4) + ":G" + str(rw + 4))
        ws.merge_cells("C" + str(rw + 5) + ":G" + str(rw + 5))
        ws.merge_cells("C" + str(rw + 6) + ":G" + str(rw + 6))

        ws.cell(row=rw + 3, column=3, value=fixed_info["document_date"])
        ws.cell(row=rw + 5, column=3, value=fixed_info["headmaster_name"])
        ws.cell(row=rw + 6, column=3, value="Okul Müdürü")
        ws.cell(row=rw + 3, column=3).alignment = alignment.Alignment(horizontal='center', vertical='center')
        ws.cell(row=rw + 5, column=3).alignment = alignment.Alignment(horizontal='center', vertical='center')
        ws.cell(row=rw + 6, column=3).alignment = alignment.Alignment(horizontal='center', vertical='center')

        try:
            sht_name = 'Sınav listesi.xlsx'
            full_address = self.save(sht_name)
            if full_address:
                wb.save(full_address)
        except PermissionError:
            tk.messagebox.showerror("Hata", "Kaydetmek istediğininz dosya açık", parent=self)

    def _create_widgets(self):
        self.std__les_btn = ttk.Button(self, text='ÇARŞAF KAYDET', command=self._create_sheet)
        self.std__les_btn.place(x=10, y=10, width=200, height=30)

        self.daily_sign_btn = ttk.Button(self, text='İMZA SİRKÜSÜ KAYDET', command=self._create_daily_signature_list)
        self.daily_sign_btn.place(x=10, y=50, width=200, height=30)

        self.daily_sign_btn = ttk.Button(self, text='SINAV LİSTESİ KAYDET', command=self._create_exam_list)
        self.daily_sign_btn.place(x=10, y=90, width=200, height=30)

        self.daily_sign_btn = ttk.Button(self, text='ÖĞRETMEN GÖREVLERİ KAYDET',
                                         command=self._create_teacher_tasks_list)
        self.daily_sign_btn.place(x=10, y=130, width=200, height=30)
