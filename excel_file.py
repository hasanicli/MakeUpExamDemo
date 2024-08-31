from openpyxl import load_workbook


class ExcelFile:
    def __init__(self, path):
        self.path = path
        self.wb = load_workbook(self.path)
        self.ws = self.wb.active

    def create_data(self):
        data = []
        if self._verify():
            lr = self.ws.max_row
            classroom_rows = self._get_classroom_rows(lr)
            lessons_list = self._get_lessons(classroom_rows)
            for lesson in lessons_list:
                exam = {"name": lesson}
                students_list = []
                for (first_rw, second_rw) in classroom_rows:
                    school_type, department, classroom = self._get_school_data(self.ws.cell(first_rw, 1).value)
                    student_rows = self._get_student_rows(first_rw, second_rw)
                    for (i, j) in student_rows:
                        number = str(self.ws.cell(i, 2).value)
                        name = str(self.ws.cell(i, 3).value).strip().replace("  ", " ")
                        for k in range(i, j):
                            if type(self.ws.cell(k, 9).value) is int:
                                lesson_name = (str(school_type) + "_" +
                                               str(self.ws.cell(k, 10).value).strip() + "_" +
                                               str(self.ws.cell(k, 9).value).strip())
                                lesson_name = self._clean_data(lesson_name)
                                if lesson == lesson_name:
                                    students_list.append({"number": number, "name": name, "school_type": school_type,
                                                          "department": department, "classroom": classroom})
                                    break
                exam["students"] = students_list
                exam["teachers"] = []
                exam["date_and_time"] = ""
                data.append(exam)
            return data
        return None

    def _verify(self):
        try:
            assert str(self.ws["B3"].value).strip() == "Öğrenci No" and \
                   str(self.ws["C3"].value).strip() == "Adı Soyadı" and \
                   str(self.ws["I3"].value).strip() == "Sınıfı" and \
                   str(self.ws["J3"].value).strip() == "Dersi", "İçerik uygun değil"
        except AssertionError:
            return False
        else:
            return True

    @staticmethod
    def _clean_data(text):
        text = text.replace("  ", " ")
        text = text.replace("/", "; ")
        text = text.replace("-", " ")
        return text

    def _get_classroom_rows(self, lr):
        """İlk sütundaki sınıf bilgilerini alır."""
        classroom_rows = [i for i in range(2, lr) if type(self.ws.cell(i, 1).value) is str] + [lr]
        return [(i, j) for index, i in enumerate(classroom_rows[:-1]) for j in classroom_rows[index + 1:index + 2]]

    def _get_lessons(self, classroom_rows):
        lessons_set = set()
        for (first_rw, second_rw) in classroom_rows:
            school_type, department, classroom = self._get_school_data(self.ws.cell(first_rw, 1).value)
            for i in range(first_rw, second_rw):
                if type(self.ws.cell(i, 9).value) is int:
                    lesson_name = (str(school_type) + "_" +
                                   str(self.ws.cell(i, 10).value).strip() + "_" +
                                   str(self.ws.cell(i, 9).value).strip())
                    lesson_name = self._clean_data(lesson_name)
                    lessons_set.add(lesson_name)
        lessons_list = list(lessons_set)
        lessons_list.sort()
        return lessons_list

    def _get_student_rows(self, first_row, second_row):
        student_rows = [i for i in range(first_row, second_row) if type(self.ws.cell(i, 2).value) is int] + [second_row]
        return [(i, j) for index, i in enumerate(student_rows) for j in student_rows[index + 1:index + 2]]

    @staticmethod
    def _get_school_data(full_info: str) -> list:
        school_type = full_info.split("-")[0].strip()
        dep = full_info.split("(")[1].split(")")[0].strip()
        classroom = full_info.split(".")[0].split("-")[1].strip() + full_info.split("/")[1].strip().split()[0].strip()
        return [school_type, dep.replace("-", " "), classroom]

    def __del__(self):
        if self.wb:
            self.wb.close()
